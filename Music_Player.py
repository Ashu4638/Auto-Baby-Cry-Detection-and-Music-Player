import openpyxl as xl
# A simple implementation of Priority Queue
# using Queue.
class PriorityQueue(object):
    def __init__(self):
        self.queue = []

    def __str__(self):
        return ' '.join([str(i) for i in self.queue])

    # for checking if the queue is empty
    def isEmpty(self):
        return len(self.queue) == 0

    # for inserting an element in the queue
    def insert(self, data, priority):
        self.queue.append([data, priority])

    # for popping an element based on Priority
    def delete(self):
        try:
            max = 0
            for i in range(len(self.queue)):
                if self.queue[i][1] > self.queue[max][1]:
                    max = i
            item = self.queue[max]
            del self.queue[max]
            return item
        except IndexError:
            print()
            exit()

def load_music():
    wb = xl.load_workbook("C:/Users/ashu_/PycharmProjects/Baby_Cry_Detection/Music.xlsx")

    sheet = wb.active
    pq = PriorityQueue()
    for row in range(2, sheet.max_row+1):
        pq.insert(sheet.cell(row, 1).value, int(sheet.cell(row, 2).value))


    return pq




